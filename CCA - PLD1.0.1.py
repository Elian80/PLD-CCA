import tkinter as tk
from tkinter import messagebox, filedialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
import time
import os
import glob
import pandas as pd
from datetime import datetime, timedelta
import traceback
import sys

# =========================
# CAMINHO BASE
# =========================
def pasta_base():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

# =========================
# VAR GLOBAL
# =========================
caminho_pld = ""
CONFIG_ARQUIVO = os.path.join(pasta_base(), "config.txt")
LOG_ARQUIVO = os.path.join(pasta_base(), "log_erros.txt")

# =========================
# LOG DE ERROS
# =========================
def registrar_log(erro):
    with open(LOG_ARQUIVO, "a", encoding="utf-8") as f:
        f.write("\n" + "=" * 80 + "\n")
        f.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + "\n")
        f.write(str(erro) + "\n")
        f.write(traceback.format_exc() + "\n")

# =========================
# SALVAR / CARREGAR CAMINHO
# =========================
def salvar_caminho(caminho):
    with open(CONFIG_ARQUIVO, "w", encoding="utf-8") as f:
        f.write(caminho)

def carregar_caminho():
    if os.path.exists(CONFIG_ARQUIVO):
        with open(CONFIG_ARQUIVO, "r", encoding="utf-8") as f:
            return f.read().strip()
    return ""

# =========================
# DATA ESCOLHIDA
# =========================
def obter_data_escolhida():
    if opcao_data.get() == "anterior":
        return datetime.now() - timedelta(days=1)
    return datetime.now()

# =========================
# SELECIONAR ARQUIVO
# =========================
def selecionar_arquivo():
    global caminho_pld

    caminho = filedialog.askopenfilename(
        title="Selecione o arquivo PLD",
        filetypes=[("Excel", "*.xlsx *.xls")]
    )

    if caminho:
        caminho_pld = caminho
        label_arquivo.config(text=os.path.basename(caminho))
        salvar_caminho(caminho)

# =========================
# COPIAR VALORES
# =========================
def copiar_valores():
    texto = text_box.get("1.0", tk.END)
    janela.clipboard_clear()
    janela.clipboard_append(texto)
    messagebox.showinfo("Copiado", "Valores copiados!")

# =========================
# CÓDIGO 1 - CCEE / SUL
# =========================
def executar_pld_hoje():
    driver = None

    try:
        text_box.delete("1.0", tk.END)
        text_box.insert(tk.END, "Iniciando navegador...\n")
        janela.update()

        driver = webdriver.Chrome()
        driver.get("https://www.ccee.org.br/en/web/guest/precos/painel-precos")
        driver.maximize_window()

        text_box.insert(tk.END, "Aguardando página carregar...\n")
        janela.update()
        time.sleep(6)

        try:
            driver.find_element(By.XPATH, "//button[contains(text(), 'Aceitar')]").click()
        except:
            pass

        time.sleep(2)

        try:
            dropdown = driver.find_element(By.TAG_NAME, "select")
            Select(dropdown).select_by_value("HORARIO")
        except:
            driver.execute_script(
                "arguments[0].click();",
                driver.find_element(By.XPATH, "//option[@value='HORARIO']")
            )

        data_escolhida = obter_data_escolhida()
        data_texto = data_escolhida.strftime("%d/%m/%Y")

        driver.execute_script(
            "arguments[0].value = arguments[1];",
            driver.find_element(By.ID, "inputInitialDate"),
            data_texto
        )

        try:
            campo_ate = driver.find_element(By.ID, "inputFinalDate")
        except:
            campo_ate = driver.find_element(By.XPATH, "(//input[@placeholder='dd/mm/aaaa'])[2]")

        driver.execute_script("arguments[0].value = arguments[1];", campo_ate, data_texto)

        text_box.insert(tk.END, f"Buscando dados de {data_texto}...\n")
        janela.update()
        time.sleep(1)

        driver.execute_script(
            "arguments[0].click();",
            driver.find_element(By.XPATH, "//button[contains(text(), 'Gerar Arquivo')]")
        )

        text_box.insert(tk.END, "Aguardando download...\n")
        janela.update()
        time.sleep(15)

        caminho_downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        padrao_xlsx = os.path.join(caminho_downloads, "preco_horario*.xlsx")
        padrao_csv = os.path.join(caminho_downloads, "preco_horario*.csv")

        arquivos = glob.glob(padrao_xlsx) + glob.glob(padrao_csv)

        if not arquivos:
            raise Exception("Arquivo baixado não foi encontrado na pasta Downloads.")

        arquivo_recente = max(arquivos, key=os.path.getmtime)

        text_box.insert(tk.END, f"Arquivo encontrado: {os.path.basename(arquivo_recente)}\n")
        janela.update()

        driver.quit()
        driver = None

        if arquivo_recente.lower().endswith(".csv"):
            df = pd.read_csv(arquivo_recente, sep=None, engine="python")
        else:
            df = pd.read_excel(arquivo_recente)

        coluna_filtro = "Submercado"

        if coluna_filtro not in df.columns:
            raise Exception(f"Coluna '{coluna_filtro}' não encontrada no arquivo.")

        df_sul = df[df[coluna_filtro].astype(str).str.contains("SUL", case=False, na=False)]

        if df_sul.empty:
            raise Exception("Nenhum dado do submercado SUL foi encontrado.")

        valores = df_sul.iloc[:, 2].dropna().tolist()

        text_box.delete("1.0", tk.END)
        for v in valores:
            text_box.insert(tk.END, f"{v}\n")

        messagebox.showinfo("Sucesso", f"Valores do SUL carregados para {data_texto}!")

    except Exception as e:
        registrar_log(e)
        messagebox.showerror("Erro", f"Falha no processamento:\n{e}\n\nVeja o arquivo log_erros.txt ao lado do executável.")

    finally:
        if driver:
            driver.quit()

# =========================
# CÓDIGO 2 - PROCESSAR PLD
# =========================
def processar_pld():
    global caminho_pld

    if not caminho_pld:
        messagebox.showwarning("Aviso", "Selecione o arquivo primeiro.")
        return

    try:
        texto = text_box.get("1.0", tk.END).strip()

        if not texto:
            messagebox.showwarning("Aviso", "Não há valores do SUL na tela.")
            return

        valores = []
        for linha in texto.splitlines():
            linha = linha.strip()
            if linha:
                try:
                    valores.append(float(str(linha).replace(",", ".")))
                except:
                    pass

        if not valores:
            messagebox.showwarning("Aviso", "Nenhum valor numérico válido encontrado.")
            return

        wb = load_workbook(caminho_pld)
        ws = wb.active

        data_escolhida = obter_data_escolhida().date()
        alterados = 0
        indice_valor = 0

        linhas_data = []
        linhas_preenchidas = 0

        for row in range(2, ws.max_row + 1):
            valor_a = ws[f"A{row}"].value

            if not valor_a:
                continue

            try:
                data_convertida = pd.to_datetime(valor_a, errors="coerce")

                if pd.isna(data_convertida):
                    continue

                if data_convertida.date() == data_escolhida:
                    linhas_data.append(row)

                    valor_d = ws[f"D{row}"].value
                    if valor_d not in (None, ""):
                        linhas_preenchidas += 1

            except:
                continue

        if not linhas_data:
            wb.close()
            messagebox.showwarning("Aviso", "Nenhuma linha com a data selecionada foi encontrada.")
            return

        if linhas_preenchidas == len(linhas_data):
            wb.close()
            messagebox.showinfo("Aviso", "Os valores dessa data já foram preenchidos.")
            return

        for row in linhas_data:
            if indice_valor < len(valores):
                ws[f"D{row}"] = valores[indice_valor]
                indice_valor += 1
                alterados += 1

        wb.save(caminho_pld)
        wb.close()

        os.startfile(caminho_pld)

        messagebox.showinfo("Sucesso", f"{alterados} linhas preenchidas com os valores do SUL.")

    except Exception as e:
        registrar_log(e)
        messagebox.showerror("Erro", f"{e}\n\nVeja o arquivo log_erros.txt ao lado do executável.")

# =========================
# INTERFACE
# =========================
janela = tk.Tk()
janela.title("Monitor PLD - SUL")
janela.geometry("420x520")

tk.Label(janela, text="Arquivo PLD:", font=("Arial", 10)).pack(pady=5)

tk.Button(janela, text="Procurar", command=selecionar_arquivo).pack()

label_arquivo = tk.Label(janela, text="Nenhum arquivo selecionado", fg="gray")
label_arquivo.pack(pady=5)

caminho_salvo = carregar_caminho()
if caminho_salvo and os.path.exists(caminho_salvo):
    caminho_pld = caminho_salvo
    label_arquivo.config(text=os.path.basename(caminho_salvo))

tk.Label(janela, text="Escolha a data:", font=("Arial", 10, "bold")).pack(pady=5)

opcao_data = tk.StringVar(value="atual")

frame_opcoes = tk.Frame(janela)
frame_opcoes.pack()

tk.Radiobutton(frame_opcoes, text="Dia atual", variable=opcao_data, value="atual").pack(side="left", padx=10)
tk.Radiobutton(frame_opcoes, text="Dia anterior", variable=opcao_data, value="anterior").pack(side="left", padx=10)

tk.Button(
    janela,
    text="BUSCAR VALORES PLD-SUL",
    command=executar_pld_hoje,
    bg="#28a745",
    fg="white",
    font=("Arial", 10, "bold"),
    padx=20,
    pady=10
).pack(pady=10)

text_box = tk.Text(janela, height=10, width=40)
text_box.pack(pady=10)

tk.Button(
    janela,
    text="COPIAR VALORES",
    command=copiar_valores,
    bg="blue",
    fg="white"
).pack(pady=5)

tk.Button(
    janela,
    text="PREENCHER PLD",
    command=processar_pld,
    bg="orange",
    fg="white",
    font=("Arial", 10, "bold"),
    padx=20,
    pady=10
).pack(pady=10)

janela.mainloop()